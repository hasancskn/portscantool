import os
import socket
import threading
import time
import ipaddress
import subprocess
import platform
from datetime import datetime, timedelta
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile
import json
from sqlalchemy.orm import Session
from models import get_db, create_tables, Scan, ScanResult, HostInfo, Alert, AuditLog, ScheduledScan, AlertRule, User
from scanner import AdvancedPortScanner
from scheduler import scheduler
from alert_service import AlertService
from config import Config
import plotly.graph_objs as go
import plotly.utils
import pandas as pd
import logging
from sqlalchemy import func

# Logging setup
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config.from_object(Config)

# Flask-Login setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Bu sayfaya erişmek için giriş yapmalısınız.'
login_manager.login_message_category = 'info'

@login_manager.user_loader
def load_user(user_id):
    db = next(get_db())
    return db.query(User).get(int(user_id))

# Veritabanı tablolarını oluştur
create_tables()

# Scheduler'ı başlat
scheduler.start()

class IPScanner:
    def __init__(self, network_range, timeout=1):
        self.network_range = network_range
        self.timeout = timeout
        self.active_ips = []
        self.scanned_ips = 0
        self.total_ips = 0
        self.lock = threading.Lock()
        
    def ping_host_socket(self, ip):
        """Socket ile ping benzeri kontrol"""
        try:
            # TCP bağlantısı ile kontrol (port 80 veya 443)
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(self.timeout)
            
            # Önce port 80'i dene
            result = sock.connect_ex((str(ip), 80))
            if result == 0:
                sock.close()
                return True
                
            # Port 80 kapalıysa port 443'ü dene
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(self.timeout)
            result = sock.connect_ex((str(ip), 443))
            sock.close()
            
            return result == 0
            
        except Exception:
            return False
    
    def ping_host(self, ip):
        """Tek bir IP adresini ping ile kontrol eder"""
        try:
            # Önce socket ile dene
            if self.ping_host_socket(ip):
                with self.lock:
                    self.scanned_ips += 1
                    hostname = self.get_hostname(ip)
                    self.active_ips.append({
                        'ip': str(ip),
                        'hostname': hostname,
                        'status': 'Active',
                        'response_time': 'TCP Check'
                    })
                return
            
            # Platform'a göre ping komutu
            if platform.system().lower() == "windows":
                # Windows için ping komutu - daha güvenilir
                ping_cmd = ["ping", "-n", "1", "-w", str(int(self.timeout * 1000)), str(ip)]
            else:
                # Linux/Mac için ping komutu
                ping_cmd = ["ping", "-c", "1", "-W", str(int(self.timeout)), str(ip)]
            
            # Subprocess timeout'u ping timeout'undan biraz daha uzun tut
            process_timeout = self.timeout + 3
            
            result = subprocess.run(
                ping_cmd, 
                capture_output=True, 
                text=True, 
                timeout=process_timeout,
                shell=False
            )
            
            with self.lock:
                self.scanned_ips += 1
                
                # Ping başarılı mı kontrol et
                if result.returncode == 0:
                    # Ping başarılı, host aktif
                    hostname = self.get_hostname(ip)
                    response_time = self.extract_response_time(result.stdout)
                    
                    self.active_ips.append({
                        'ip': str(ip),
                        'hostname': hostname,
                        'status': 'Active',
                        'response_time': response_time
                    })
                    
        except subprocess.TimeoutExpired:
            # Ping timeout oldu
            with self.lock:
                self.scanned_ips += 1
        except Exception as e:
            # Diğer hatalar
            with self.lock:
                self.scanned_ips += 1
                print(f"Ping error for {ip}: {str(e)}")
    
    def get_hostname(self, ip):
        """IP adresinden hostname almaya çalışır"""
        try:
            # Hostname çözümleme için timeout ayarla
            socket.setdefaulttimeout(2)
            hostname = socket.gethostbyaddr(str(ip))[0]
            return hostname
        except socket.herror:
            # Hostname bulunamadı
            return "Unknown"
        except socket.timeout:
            # Timeout
            return "Timeout"
        except Exception:
            # Diğer hatalar
            return "Error"
        finally:
            # Timeout'u sıfırla
            socket.setdefaulttimeout(None)
    
    def extract_response_time(self, ping_output):
        """Ping çıktısından response time çıkarır"""
        try:
            if platform.system().lower() == "windows":
                # Windows ping formatı: "time=5ms" veya "time<1ms"
                if "time=" in ping_output:
                    time_part = ping_output.split("time=")[1].split()[0]
                    return time_part
                elif "time<" in ping_output:
                    time_part = ping_output.split("time<")[1].split()[0]
                    return f"<{time_part}"
            else:
                # Linux/Mac ping formatı: "time=5.123 ms"
                if "time=" in ping_output:
                    time_part = ping_output.split("time=")[1].split()[0]
                    return time_part + " ms"
        except:
            pass
        return "Unknown"
    
    def scan_network(self, max_threads=50):
        """Ağ taramasını başlatır"""
        try:
            # Network range'i parse et
            network = ipaddress.IPv4Network(self.network_range, strict=False)
            self.total_ips = network.num_addresses
            
            print(f"Starting network scan for {self.network_range}")
            print(f"Total IPs to scan: {self.total_ips}")
            print(f"Max threads: {max_threads}")
            
            threads = []
            
            for ip in network.hosts():
                thread = threading.Thread(target=self.ping_host, args=(ip,))
                threads.append(thread)
                thread.start()
                
                # Maksimum thread sayısını kontrol et
                if len(threads) >= max_threads:
                    for t in threads:
                        t.join()
                    threads = []
            
            # Kalan thread'leri bekle
            for t in threads:
                t.join()
            
            print(f"Scan completed. Found {len(self.active_ips)} active IPs")
            return self.active_ips, self.scanned_ips, self.total_ips
            
        except Exception as e:
            print(f"Network scan error: {str(e)}")
            raise Exception(f"Ağ tarama hatası: {str(e)}")

class PortScanner:
    def __init__(self, target_host, start_port, end_port, timeout=1):
        self.target_host = target_host
        self.start_port = start_port
        self.end_port = end_port
        self.timeout = timeout
        self.open_ports = []
        self.scanned_ports = 0
        self.total_ports = end_port - start_port + 1
        self.lock = threading.Lock()
        
    def scan_port(self, port):
        """Tek bir portu tarar"""
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(self.timeout)
            result = sock.connect_ex((self.target_host, port))
            sock.close()
            
            with self.lock:
                self.scanned_ports += 1
                if result == 0:
                    self.open_ports.append({
                        'port': port,
                        'service': self.get_service_name(port),
                        'status': 'Open'
                    })
                    
        except Exception as e:
            with self.lock:
                self.scanned_ports += 1
                
    def get_service_name(self, port):
        """Port numarasına göre servis adını döndürür"""
        common_services = {
            21: 'FTP',
            22: 'SSH',
            23: 'Telnet',
            25: 'SMTP',
            53: 'DNS',
            80: 'HTTP',
            110: 'POP3',
            143: 'IMAP',
            443: 'HTTPS',
            993: 'IMAP SSL',
            995: 'POP3 SSL',
            1433: 'MSSQL',
            3306: 'MySQL',
            3389: 'RDP',
            5432: 'PostgreSQL',
            5900: 'VNC',
            6379: 'Redis',
            8080: 'HTTP Proxy',
            8443: 'HTTPS Alt',
            27017: 'MongoDB'
        }
        return common_services.get(port, 'Unknown')
    
    def scan(self, max_threads=100):
        """Port taramasını başlatır"""
        threads = []
        
        for port in range(self.start_port, self.end_port + 1):
            thread = threading.Thread(target=self.scan_port, args=(port,))
            threads.append(thread)
            thread.start()
            
            # Maksimum thread sayısını kontrol et
            if len(threads) >= max_threads:
                for t in threads:
                    t.join()
                threads = []
        
        # Kalan thread'leri bekle
        for t in threads:
            t.join()
            
        return self.open_ports, self.scanned_ports, self.total_ports

@app.route('/')
@login_required
def index():
    """Ana sayfa"""
    return render_template('index.html')

@app.route('/scan-network', methods=['POST'])
@login_required
def scan_network():
    """Ağ taraması API endpoint'i"""
    try:
        data = request.get_json()
        network_range = data.get('network_range')
        timeout = float(data.get('timeout', 1))
        max_threads = int(data.get('max_threads', 50))
        print(f"Network scan request: {network_range}, timeout: {timeout}, threads: {max_threads}")
        try:
            ipaddress.IPv4Network(network_range, strict=False)
        except Exception:
            return jsonify({'error': 'Geçersiz ağ aralığı (örn: 192.168.1.0/24)'}), 400
        scanner = IPScanner(network_range, timeout)
        active_ips, scanned_ips, total_ips = scanner.scan_network(max_threads)
        scan_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        result = {
            'network_range': network_range,
            'scan_time': scan_time,
            'total_ips': total_ips,
            'scanned_ips': scanned_ips,
            'active_ips_count': len(active_ips),
            'active_ips': active_ips
        }
        try:
            db = next(get_db())
            from models import NetworkScan, NetworkScanResult
            network_scan = NetworkScan(
                network_range=network_range,
                active_ip_count=len(active_ips),
                scan_type='manual',
                created_at=datetime.now()
            )
            db.add(network_scan)
            db.commit()
            # Aktif IP'leri NetworkScanResult tablosuna kaydet
            for ipinfo in active_ips:
                nsr = NetworkScanResult(
                    network_scan_id=network_scan.id,
                    ip=ipinfo.get('ip'),
                    hostname=ipinfo.get('hostname'),
                    status=ipinfo.get('status'),
                    response_time=ipinfo.get('response_time')
                )
                db.add(nsr)
            db.commit()
        except Exception as e:
            print(f"NetworkScan DB kayıt hatası: {str(e)}")
        print(f"Scan completed: {len(active_ips)} active IPs found")
        return jsonify(result)
    except Exception as e:
        print(f"Network scan error: {str(e)}")
        return jsonify({'error': f'Ağ tarama hatası: {str(e)}'}), 500

@app.route('/scan', methods=['POST'])
@login_required
def scan_host():
    """Manuel port taraması"""
    try:
        data = request.get_json()
        target_host = data.get('target_host')
        start_port = int(data.get('start_port', 1))
        end_port = int(data.get('end_port', 1024))
        
        db = next(get_db())
        scanner = AdvancedPortScanner(db)
        
        # Kullanıcı bilgilerini al
        user = request.headers.get('X-User', 'anonymous')
        ip_address = request.remote_addr
        
        result = scanner.scan_host(
            target_host=target_host,
            start_port=start_port,
            end_port=end_port,
            scan_type='manual',
            user=user,
            ip_address=ip_address
        )
        
        return jsonify({
            'success': True,
            'scan_id': result['scan_id'],
            'target_host': result['target_host'],
            'total_ports': result['total_ports'],
            'scanned_ports': result['scanned_ports'],
            'open_ports_count': result['open_ports_count'],
            'scan_duration': result['scan_duration'],
            'scan_results': [
                {
                    'port': r.port,
                    'service': r.service,
                    'version': r.version,
                    'state': r.state,
                    'protocol': r.protocol,
                    'banner': r.banner
                } for r in result['scan_results']
            ] if result['scan_results'] else []
        })
        
    except Exception as e:
        logger.error(f"Scan error: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/scans')
@login_required
def get_scans():
    """Tüm taramaları listele"""
    try:
        db = next(get_db())
        scans = db.query(Scan).order_by(Scan.start_time.desc()).limit(50).all()
        
        result = []
        for scan in scans:
            result.append({
                'id': scan.id,
                'target_host': scan.target_host,
                'start_port': scan.start_port,
                'end_port': scan.end_port,
                'scan_type': scan.scan_type,
                'status': scan.status,
                'start_time': scan.start_time.isoformat(),
                'end_time': scan.end_time.isoformat() if scan.end_time else None,
                'total_ports': scan.total_ports,
                'open_ports_count': scan.open_ports_count,
                'scan_duration': scan.scan_duration
            })
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Error getting scans: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/scan/<int:scan_id>')
@login_required
def get_scan_details(scan_id):
    """Tarama detaylarını getir"""
    try:
        db = next(get_db())
        scan = db.query(Scan).filter(Scan.id == scan_id).first()
        
        if not scan:
            return jsonify({'error': 'Scan not found'}), 404
        
        # Scan sonuçları
        scan_results = db.query(ScanResult).filter(ScanResult.scan_id == scan_id).all()
        
        # Host bilgileri
        host_info = db.query(HostInfo).filter(HostInfo.scan_id == scan_id).first()
        
        # Alarmlar
        alerts = db.query(Alert).filter(Alert.scan_id == scan_id).all()
        
        result = {
            'scan': {
                'id': scan.id,
                'target_host': scan.target_host,
                'start_port': scan.start_port,
                'end_port': scan.end_port,
                'scan_type': scan.scan_type,
                'status': scan.status,
                'start_time': scan.start_time.isoformat(),
                'end_time': scan.end_time.isoformat() if scan.end_time else None,
                'total_ports': scan.total_ports,
                'open_ports_count': scan.open_ports_count,
                'scan_duration': scan.scan_duration
            },
            'scan_results': [
                {
                    'port': r.port,
                    'service': r.service,
                    'version': r.version,
                    'state': r.state,
                    'protocol': r.protocol,
                    'banner': r.banner
                } for r in scan_results
            ],
            'host_info': {
                'ip_address': host_info.ip_address,
                'hostname': host_info.hostname,
                'os_family': host_info.os_family,
                'os_version': host_info.os_version,
                'os_accuracy': host_info.os_accuracy,
                'mac_address': host_info.mac_address,
                'vendor': host_info.vendor
            } if host_info else None,
            'alerts': [
                {
                    'id': a.id,
                    'alert_type': a.alert_type,
                    'severity': a.severity,
                    'message': a.message,
                    'details': a.details,
                    'created_at': a.created_at.isoformat(),
                    'is_sent': a.is_sent
                } for a in alerts
            ]
        }
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Error getting scan details: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/scheduled-scans', methods=['GET', 'POST'])
@login_required
def scheduled_scans():
    """Zamanlanmış taramaları yönet"""
    if request.method == 'GET':
        try:
            scans = scheduler.get_scheduled_scans()
            return jsonify(scans)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    elif request.method == 'POST':
        # Monitor kullanıcıları zamanlanmış tarama ekleyemez
        if current_user.role == 'monitor':
            return jsonify({'error': 'Monitor kullanıcıları zamanlanmış tarama ekleyemez'}), 403
        
        try:
            data = request.get_json()
            name = data.get('name')
            scan_type = data.get('scan_type')
            target_hosts = data.get('target_hosts', [])
            port_range = data.get('port_range', '1-1024')
            network_range = data.get('network_range', '')
            schedule_interval = int(data.get('schedule_interval', 3600))
            success = scheduler.add_scheduled_scan(name, scan_type, target_hosts, port_range, network_range, schedule_interval)
            if success:
                return jsonify({'success': True, 'message': 'Scheduled scan added successfully'})
            else:
                return jsonify({'error': 'Failed to add scheduled scan'}), 500
        except Exception as e:
            return jsonify({'error': str(e)}), 500

@app.route('/scheduled-scans/<int:scan_id>', methods=['DELETE'])
@login_required
def remove_scheduled_scan(scan_id):
    """Zamanlanmış taramayı kaldır"""
    # Monitor kullanıcıları zamanlanmış tarama silemez
    if current_user.role == 'monitor':
        return jsonify({'error': 'Monitor kullanıcıları zamanlanmış tarama silemez'}), 403
    
    try:
        success = scheduler.remove_scheduled_scan(scan_id)
        
        if success:
            return jsonify({'success': True, 'message': 'Scheduled scan removed successfully'})
        else:
            return jsonify({'error': 'Failed to remove scheduled scan'}), 500
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/alerts')
@login_required
def get_alerts():
    """Alarmları listele"""
    try:
        db = next(get_db())
        alerts = db.query(Alert).order_by(Alert.created_at.desc()).limit(100).all()
        
        result = []
        for alert in alerts:
            result.append({
                'id': alert.id,
                'scan_id': alert.scan_id,
                'alert_type': alert.alert_type,
                'severity': alert.severity,
                'message': alert.message,
                'details': alert.details,
                'created_at': alert.created_at.isoformat(),
                'is_sent': alert.is_sent,
                'sent_time': alert.sent_time.isoformat() if alert.sent_time else None
            })
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Error getting alerts: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/test-alert', methods=['POST'])
@login_required
def test_alert():
    """Test alarmı gönder"""
    try:
        alert_service = AlertService()
        alert_service.send_test_alert()
        return jsonify({'success': True, 'message': 'Test alert sent successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard')
@login_required
def dashboard():
    """Dashboard sayfası"""
    return render_template('dashboard.html')

@app.route('/api/dashboard-stats')
@login_required
def dashboard_stats():
    """Dashboard istatistikleri"""
    try:
        db = next(get_db())
        
        # Son 30 günün verilerini al
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        
        # Toplam tarama sayısı
        total_scans = db.query(Scan).count()
        
        # Son 30 günde yapılan taramalar
        recent_scans = db.query(Scan).filter(Scan.start_time >= thirty_days_ago).count()
        
        # Toplam açık port sayısı
        total_open_ports = db.query(ScanResult).filter(ScanResult.state == 'open').count()
        
        # Toplam alarm sayısı
        total_alerts = db.query(Alert).count()
        
        # Son 30 günde oluşan alarmlar
        recent_alerts = db.query(Alert).filter(Alert.created_at >= thirty_days_ago).count()
        
        # Günlük tarama grafiği için veri
        daily_scans = db.query(Scan).filter(Scan.start_time >= thirty_days_ago).all()
        scan_dates = {}
        for scan in daily_scans:
            date = scan.start_time.date().isoformat()
            scan_dates[date] = scan_dates.get(date, 0) + 1
        
        # Port dağılımı
        port_distribution = db.query(ScanResult.service, func.count(ScanResult.id)).filter(
            ScanResult.state == 'open'
        ).group_by(ScanResult.service).order_by(func.count(ScanResult.id).desc()).limit(10).all()
        
        return jsonify({
            'total_scans': total_scans,
            'recent_scans': recent_scans,
            'total_open_ports': total_open_ports,
            'total_alerts': total_alerts,
            'recent_alerts': recent_alerts,
            'daily_scans': scan_dates,
            'port_distribution': [{'service': service, 'count': count} for service, count in port_distribution]
        })
        
    except Exception as e:
        logger.error(f"Error getting dashboard stats: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/api/network-scan-stats')
@login_required
def network_scan_stats():
    """Ağ tarama dashboardu için istatistikler"""
    try:
        db = next(get_db())
        from models import NetworkScan
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)

        # Son 30 günün tüm ağ taramaları (manuel + scheduled)
        network_scans = db.query(NetworkScan).filter(
            NetworkScan.created_at >= thirty_days_ago
        ).order_by(NetworkScan.created_at.desc()).all()

        # Toplam ağ taraması
        total_network_scans = db.query(NetworkScan).count()

        # Toplam aktif IP
        total_active_ips = db.query(func.sum(NetworkScan.active_ip_count)).scalar() or 0

        # Günlük ağ taramaları (son 30 gün)
        daily_network_scans = {}
        for scan in network_scans:
            date = scan.created_at.date().isoformat()
            daily_network_scans[date] = daily_network_scans.get(date, 0) + 1

        # Son 10 ağ taraması
        last_network_scans = []
        for scan in network_scans[:10]:
            last_network_scans.append({
                'network_range': scan.network_range,
                'active_ip_count': scan.active_ip_count,
                'scan_type': scan.scan_type,
                'created_at': scan.created_at.isoformat()
            })

        return jsonify({
            'total_network_scans': total_network_scans,
            'total_active_ips': total_active_ips,
            'daily_network_scans': daily_network_scans,
            'last_network_scans': last_network_scans
        })
    except Exception as e:
        logger.error(f"Error getting network scan stats: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/api/network-scans')
@login_required
def api_network_scans():
    """Tüm ağ taramalarını filtreleme, arama ve pagination ile JSON döner"""
    try:
        db = next(get_db())
        from models import NetworkScan
        query = db.query(NetworkScan)
        # Filtreler
        scan_type = request.args.get('scan_type')
        network_range = request.args.get('network_range')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        if scan_type:
            query = query.filter(NetworkScan.scan_type == scan_type)
        if network_range:
            query = query.filter(NetworkScan.network_range.ilike(f"%{network_range}%"))
        if date_from:
            query = query.filter(NetworkScan.created_at >= date_from)
        if date_to:
            query = query.filter(NetworkScan.created_at <= date_to)
        # Pagination
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 10))
        total = query.count()
        scans = query.order_by(NetworkScan.created_at.desc()).offset((page-1)*page_size).limit(page_size).all()
        result = []
        for scan in scans:
            result.append({
                'id': scan.id,
                'network_range': scan.network_range,
                'active_ip_count': scan.active_ip_count,
                'scan_type': scan.scan_type,
                'created_at': scan.created_at.isoformat()
            })
        return jsonify({'items': result, 'total': total})
    except Exception as e:
        logger.error(f"Error getting network scans: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/api/port-scans')
@login_required
def api_port_scans():
    """Tüm port taramalarını filtreleme, arama ve pagination ile JSON döner"""
    try:
        db = next(get_db())
        from models import Scan
        query = db.query(Scan)
        scan_type = request.args.get('scan_type')
        target_host = request.args.get('target_host')
        port_range = request.args.get('port_range')
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        if scan_type:
            query = query.filter(Scan.scan_type == scan_type)
        if target_host:
            query = query.filter(Scan.target_host.ilike(f"%{target_host}%"))
        if port_range:
            try:
                start, end = map(int, port_range.split('-'))
                query = query.filter(Scan.start_port >= start, Scan.end_port <= end)
            except:
                pass
        if date_from:
            query = query.filter(Scan.start_time >= date_from)
        if date_to:
            query = query.filter(Scan.start_time <= date_to)
        # Pagination
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 10))
        total = query.count()
        scans = query.order_by(Scan.start_time.desc()).offset((page-1)*page_size).limit(page_size).all()
        result = []
        for scan in scans:
            result.append({
                'id': scan.id,
                'target_host': scan.target_host,
                'start_port': scan.start_port,
                'end_port': scan.end_port,
                'scan_type': scan.scan_type,
                'status': scan.status,
                'open_ports_count': scan.open_ports_count,
                'start_time': scan.start_time.isoformat() if scan.start_time else '',
                'end_time': scan.end_time.isoformat() if scan.end_time else ''
            })
        return jsonify({'items': result, 'total': total})
    except Exception as e:
        logger.error(f"Error getting port scans: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/export-excel', methods=['POST'])
@login_required
def export_excel():
    """Excel export (eski fonksiyonalite korundu)"""
    try:
        data = request.get_json()
        scan_type = data.get('scan_type')
        scan_results = data.get('scan_results')
        
        import openpyxl
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if scan_type == 'port':
            ws.title = "Port Scan Results"
            headers = ['Port', 'Service', 'Version', 'State', 'Protocol', 'Banner']
            ws.append(headers)
            
            for result in scan_results.get('scan_results', []):
                ws.append([
                    result.get('port'),
                    result.get('service'),
                    result.get('version'),
                    result.get('state'),
                    result.get('protocol'),
                    result.get('banner')
                ])
        
        elif scan_type == 'network':
            ws.title = "Network Scan Results"
            headers = ['IP Address', 'Hostname', 'Status', 'Response Time']
            ws.append(headers)
            
            for result in scan_results.get('active_ips', []):
                ws.append([
                    result.get('ip'),
                    result.get('hostname'),
                    result.get('status'),
                    result.get('response_time')
                ])
        
        # Excel dosyasını kaydet
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'scan_{scan_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Excel export error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/export-network-scans', methods=['POST'])
@login_required
def export_network_scans():
    """Filtreli ağ taramalarını Excel olarak döner"""
    try:
        data = request.get_json() or {}
        scan_type = data.get('scan_type')
        network_range = data.get('network_range')
        date_from = data.get('date_from')
        date_to = data.get('date_to')
        
        db = next(get_db())
        from models import NetworkScan
        query = db.query(NetworkScan)
        if scan_type:
            query = query.filter(NetworkScan.scan_type == scan_type)
        if network_range:
            query = query.filter(NetworkScan.network_range.ilike(f"%{network_range}%"))
        if date_from:
            query = query.filter(NetworkScan.created_at >= date_from)
        if date_to:
            query = query.filter(NetworkScan.created_at <= date_to)
        scans = query.order_by(NetworkScan.created_at.desc()).all()
        
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Network Scans"
        headers = ['ID', 'Ağ Aralığı', 'Aktif IP', 'Tarama Tipi', 'Tarih']
        ws.append(headers)
        for scan in scans:
            ws.append([
                scan.id,
                scan.network_range,
                scan.active_ip_count,
                scan.scan_type,
                scan.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'network_scans_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        logger.error(f"Excel export error: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/export-network-scan/<int:scan_id>')
@login_required
def export_single_network_scan(scan_id):
    """Belirli bir ağ taramasının detaylarını (aktif IP'ler) Excel olarak döner"""
    try:
        db = next(get_db())
        from models import NetworkScan, NetworkScanResult
        scan = db.query(NetworkScan).filter(NetworkScan.id == scan_id).first()
        if not scan:
            return jsonify({'error': 'Kayıt bulunamadı'}), 404
        results = db.query(NetworkScanResult).filter(NetworkScanResult.network_scan_id == scan_id).all()
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Network Scan Detail"
        headers = ['Ağ Aralığı', 'Aktif IP Sayısı', 'Tarama Tipi', 'Tarih']
        ws.append(headers)
        ws.append([
            scan.network_range,
            scan.active_ip_count,
            scan.scan_type,
            scan.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ])
        ws.append([])
        ws.append(['Aktif IP', 'Hostname', 'Status', 'Response Time'])
        for r in results:
            ws.append([
                r.ip,
                r.hostname,
                r.status,
                r.response_time
            ])
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'network_scan_{scan_id}_{scan.created_at.strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        logger.error(f"Excel export error: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/export-port-scans', methods=['POST'])
@login_required
def export_port_scans():
    """Filtreli port taramalarını Excel olarak döner"""
    try:
        data = request.get_json() or {}
        scan_type = data.get('scan_type')
        target_host = data.get('target_host')
        port_range = data.get('port_range')
        date_from = data.get('date_from')
        date_to = data.get('date_to')
        db = next(get_db())
        from models import Scan
        query = db.query(Scan)
        if scan_type:
            query = query.filter(Scan.scan_type == scan_type)
        if target_host:
            query = query.filter(Scan.target_host.ilike(f"%{target_host}%"))
        if port_range:
            try:
                start, end = map(int, port_range.split('-'))
                query = query.filter(Scan.start_port >= start, Scan.end_port <= end)
            except:
                pass
        if date_from:
            query = query.filter(Scan.start_time >= date_from)
        if date_to:
            query = query.filter(Scan.start_time <= date_to)
        scans = query.order_by(Scan.start_time.desc()).all()
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Port Scans"
        headers = ['ID', 'Hedef Host', 'Port Aralığı', 'Tarama Tipi', 'Durum', 'Açık Port', 'Başlangıç', 'Bitiş']
        ws.append(headers)
        for scan in scans:
            ws.append([
                scan.id,
                scan.target_host,
                f"{scan.start_port}-{scan.end_port}",
                scan.scan_type,
                scan.status,
                scan.open_ports_count,
                scan.start_time.strftime('%Y-%m-%d %H:%M:%S') if scan.start_time else '',
                scan.end_time.strftime('%Y-%m-%d %H:%M:%S') if scan.end_time else ''
            ])
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'port_scans_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        logger.error(f"Excel export error: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/export-port-scan/<int:scan_id>')
@login_required
def export_single_port_scan(scan_id):
    """Belirli bir port taramasının detaylarını (açık portlar) Excel olarak döner"""
    try:
        db = next(get_db())
        from models import Scan, ScanResult
        scan = db.query(Scan).filter(Scan.id == scan_id).first()
        if not scan:
            return jsonify({'error': 'Kayıt bulunamadı'}), 404
        results = db.query(ScanResult).filter(ScanResult.scan_id == scan_id).all()
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Port Scan Detail"
        headers = ['Port', 'Service', 'Version', 'State', 'Protocol', 'Banner']
        ws.append(headers)
        for r in results:
            ws.append([
                r.port,
                r.service,
                r.version,
                r.state,
                r.protocol,
                r.banner
            ])
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'port_scan_{scan_id}_{scan.start_time.strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        logger.error(f"Excel export error: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/scheduler')
@login_required
def scheduler_page():
    return render_template('scheduler.html')

@app.route('/network-scans')
@login_required
def network_scans_page():
    return render_template('network_scans.html')

@app.route('/alert-rules')
@login_required
def alert_rules_page():
    return render_template('alert_rules.html')

@app.route('/api/alert-rules', methods=['GET'])
@login_required
def get_alert_rules():
    try:
        db = next(get_db())
        rules = db.query(AlertRule).all()
        result = []
        for rule in rules:
            cond = rule.get_condition_dict()
            result.append({
                'id': rule.id,
                'name': rule.name,
                'rule_type': rule.rule_type,
                'condition': cond,
                'field': cond.get('field'),
                'operator': cond.get('operator'),
                'value': cond.get('value'),
                'severity': rule.severity,
                'action': rule.action,
                'is_active': rule.is_active,
                'created_at': rule.created_at.isoformat() if rule.created_at else None
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/api/alert-rules', methods=['POST'])
@login_required
def add_alert_rule():
    # Monitor kullanıcıları alarm kuralı ekleyemez
    if current_user.role == 'monitor':
        return jsonify({'error': 'Monitor kullanıcıları alarm kuralı ekleyemez'}), 403
    
    try:
        data = request.get_json()
        name = data.get('name')
        rule_type = data.get('rule_type')
        condition = data.get('condition')
        severity = data.get('severity', 'medium')
        action = data.get('action', 'email')
        is_active = data.get('is_active', True)
        if not name or not rule_type or not condition:
            return jsonify({'error': 'Eksik alanlar var'}), 400
        db = next(get_db())
        rule = AlertRule(
            name=name,
            rule_type=rule_type,
            condition=json.dumps(condition),
            severity=severity,
            action=action,
            is_active=is_active
        )
        db.add(rule)
        db.commit()
        return jsonify({'success': True, 'id': rule.id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/api/alert-rules/<int:rule_id>', methods=['PUT'])
@login_required
def update_alert_rule(rule_id):
    # Monitor kullanıcıları alarm kuralı düzenleyemez
    if current_user.role == 'monitor':
        return jsonify({'error': 'Monitor kullanıcıları alarm kuralı düzenleyemez'}), 403
    
    try:
        data = request.get_json()
        db = next(get_db())
        rule = db.query(AlertRule).filter(AlertRule.id == rule_id).first()
        if not rule:
            return jsonify({'error': 'Kural bulunamadı'}), 404
        rule.name = data.get('name', rule.name)
        rule.rule_type = data.get('rule_type', rule.rule_type)
        if 'condition' in data and data['condition']:
            rule.condition = json.dumps(data['condition'])
        rule.severity = data.get('severity', rule.severity)
        rule.action = data.get('action', rule.action)
        rule.is_active = data.get('is_active', rule.is_active)
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()

@app.route('/api/alert-rules/<int:rule_id>', methods=['DELETE'])
@login_required
def delete_alert_rule(rule_id):
    # Monitor kullanıcıları alarm kuralı silemez
    if current_user.role == 'monitor':
        return jsonify({'error': 'Monitor kullanıcıları alarm kuralı silemez'}), 403
    
    try:
        db = next(get_db())
        rule = db.query(AlertRule).filter(AlertRule.id == rule_id).first()
        
        if not rule:
            return jsonify({'success': False, 'error': 'Kural bulunamadı'}), 404
        
        db.delete(rule)
        db.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.rollback()
        logger.error(f"Error deleting alert rule: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

# Authentication routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if not username or not password:
            flash('Kullanıcı adı ve şifre gereklidir.', 'error')
            return render_template('login.html')
        
        db = next(get_db())
        user = db.query(User).filter(User.username == username).first()
        
        if user and check_password_hash(user.password_hash, password):
            if not user.is_active:
                flash('Hesabınız aktif değil.', 'error')
                return render_template('login.html')
            
            login_user(user)
            user.last_login = datetime.utcnow()
            db.commit()
            
            next_page = request.args.get('next')
            if not next_page or not next_page.startswith('/'):
                next_page = url_for('index')
            
            flash(f'Hoş geldiniz, {user.username}!', 'success')
            return redirect(next_page)
        else:
            flash('Geçersiz kullanıcı adı veya şifre.', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Başarıyla çıkış yaptınız.', 'info')
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        if not all([username, email, password, confirm_password]):
            flash('Tüm alanlar gereklidir.', 'error')
            return render_template('register.html')
        
        if password != confirm_password:
            flash('Şifreler eşleşmiyor.', 'error')
            return render_template('register.html')
        
        if len(password) < 6:
            flash('Şifre en az 6 karakter olmalıdır.', 'error')
            return render_template('register.html')
        
        db = next(get_db())
        
        # Kullanıcı adı veya email zaten var mı kontrol et
        existing_user = db.query(User).filter(
            (User.username == username) | (User.email == email)
        ).first()
        
        if existing_user:
            flash('Bu kullanıcı adı veya email zaten kullanılıyor.', 'error')
            return render_template('register.html')
        
        # İlk kullanıcıyı admin yap
        is_admin = db.query(User).count() == 0
        
        new_user = User(
            username=username,
            email=email,
            password_hash=generate_password_hash(password),
            is_admin=is_admin
        )
        
        db.add(new_user)
        db.commit()
        
        flash('Hesabınız başarıyla oluşturuldu! Şimdi giriş yapabilirsiniz.', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html')

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        if not all([current_password, new_password, confirm_password]):
            flash('Tüm alanlar gereklidir.', 'error')
            return render_template('change_password.html')
        
        # Mevcut şifreyi kontrol et
        if not check_password_hash(current_user.password_hash, current_password):
            flash('Mevcut şifre yanlış.', 'error')
            return render_template('change_password.html')
        
        # Yeni şifreleri kontrol et
        if new_password != confirm_password:
            flash('Yeni şifreler eşleşmiyor.', 'error')
            return render_template('change_password.html')
        
        if len(new_password) < 6:
            flash('Yeni şifre en az 6 karakter olmalıdır.', 'error')
            return render_template('change_password.html')
        
        # Şifreyi güncelle
        db = next(get_db())
        user = db.query(User).filter(User.id == current_user.id).first()
        user.password_hash = generate_password_hash(new_password)
        db.commit()
        
        flash('Şifreniz başarıyla değiştirildi!', 'success')
        return redirect(url_for('index'))
    
    return render_template('change_password.html')

@app.route('/user-management')
@login_required
def user_management():
    if not current_user.is_admin:
        flash('Bu sayfaya erişim yetkiniz yok.', 'error')
        return redirect(url_for('index'))
    
    db = next(get_db())
    users = db.query(User).order_by(User.created_at.desc()).all()
    return render_template('user_management.html', users=users)

@app.route('/create-user', methods=['GET', 'POST'])
@login_required
def create_user():
    if not current_user.is_admin:
        flash('Bu sayfaya erişim yetkiniz yok.', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        role = request.form.get('role', 'monitor')
        
        if not all([username, email, password, confirm_password]):
            flash('Tüm alanlar gereklidir.', 'error')
            return render_template('create_user.html')
        
        if password != confirm_password:
            flash('Şifreler eşleşmiyor.', 'error')
            return render_template('create_user.html')
        
        if len(password) < 6:
            flash('Şifre en az 6 karakter olmalıdır.', 'error')
            return render_template('create_user.html')
        
        db = next(get_db())
        
        # Kullanıcı adı veya email zaten var mı kontrol et
        existing_user = db.query(User).filter(
            (User.username == username) | (User.email == email)
        ).first()
        
        if existing_user:
            flash('Bu kullanıcı adı veya email zaten kullanılıyor.', 'error')
            return render_template('create_user.html')
        
        # Role'e göre is_admin belirle
        is_admin = role == 'admin'
        
        new_user = User(
            username=username,
            email=email,
            password_hash=generate_password_hash(password),
            role=role,
            is_admin=is_admin,
            created_by=current_user.id
        )
        
        db.add(new_user)
        db.commit()
        
        flash(f'{username} kullanıcısı başarıyla oluşturuldu!', 'success')
        return redirect(url_for('user_management'))
    
    return render_template('create_user.html')

@app.route('/edit-user/<int:user_id>', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):
    if not current_user.is_admin:
        flash('Bu sayfaya erişim yetkiniz yok.', 'error')
        return redirect(url_for('index'))
    
    db = next(get_db())
    user = db.query(User).filter(User.id == user_id).first()
    
    if not user:
        flash('Kullanıcı bulunamadı.', 'error')
        return redirect(url_for('user_management'))
    
    if request.method == 'POST':
        email = request.form.get('email')
        role = request.form.get('role', 'monitor')
        is_active = request.form.get('is_active') == 'on'
        
        if not email:
            flash('Email alanı gereklidir.', 'error')
            return render_template('edit_user.html', user=user)
        
        # Email değişikliği varsa kontrol et
        if email != user.email:
            existing_user = db.query(User).filter(User.email == email).first()
            if existing_user:
                flash('Bu email adresi zaten kullanılıyor.', 'error')
                return render_template('edit_user.html', user=user)
        
        # Role'e göre is_admin belirle
        is_admin = role == 'admin'
        
        user.email = email
        user.role = role
        user.is_admin = is_admin
        user.is_active = is_active
        
        db.commit()
        
        flash(f'{user.username} kullanıcısı başarıyla güncellendi!', 'success')
        return redirect(url_for('user_management'))
    
    return render_template('edit_user.html', user=user)

@app.route('/delete-user/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    if not current_user.is_admin:
        return jsonify({'success': False, 'error': 'Yetkisiz erişim'}), 403
    
    if current_user.id == user_id:
        return jsonify({'success': False, 'error': 'Kendinizi silemezsiniz'}), 400
    
    db = next(get_db())
    user = db.query(User).filter(User.id == user_id).first()
    
    if not user:
        return jsonify({'success': False, 'error': 'Kullanıcı bulunamadı'}), 404
    
    username = user.username
    db.delete(user)
    db.commit()
    
    return jsonify({'success': True, 'message': f'{username} kullanıcısı silindi'})

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000) 